import re
import sys
from pathlib import Path

import pandas as pd
import torch
import torch.nn as nn
from PIL import Image
from torchvision import models
from openpyxl import load_workbook

import src.rating as rt


# ======================================================================
# User-configurable constants
# ======================================================================
DATA_ROOT       = Path(r"C:\data\Carreteras\HD15_8m")
LABELS_XLSX     = Path(r"D:\SOTER\SOTER\scripts\rating\data_inspection\labels.xlsx")
DATASETS_DIR    = Path(r"D:\SOTER\SOTER\scripts\rating\datasets")
EXPERIMENT_NAME = "E12"

FINAL_SHEET = "FINAL"
IMG_WIDTH   = 54
IMG_HEIGHT  = 202
# ======================================================================


def build_inference_model(backbone: str, use_ca: bool, reduction_ratio: int = 4) -> nn.Module:
    """Reconstruct model architecture for inference (no pretrained weights)."""
    if backbone == "resnet18":
        m = models.resnet18(weights=None);    m.fc = nn.Linear(m.fc.in_features, 1)
    elif backbone == "resnet50":
        m = models.resnet50(weights=None);    m.fc = nn.Linear(m.fc.in_features, 1)
    elif backbone == "resnet152":
        m = models.resnet152(weights=None);   m.fc = nn.Linear(m.fc.in_features, 1)
    elif backbone == "densenet121":
        m = models.densenet121(weights=None); m.classifier = nn.Linear(m.classifier.in_features, 1)
    else:
        raise ValueError(f"Unknown backbone: {backbone}")

    if use_ca:
        if backbone in rt._RESNET_CHANNELS:
            m = rt.ResNetWithHVCA(m, backbone, reduction_ratio)
        elif backbone in rt._DENSENET_CHANNELS:
            m = rt.DenseNetWithHVCA(m, backbone, reduction_ratio)
    return m


def parse_log_config(log_path: Path) -> dict:
    """Extract backbone and attention settings from training_log.txt."""
    cfg = {"backbone": "resnet152", "use_channel_attention": False, "attention_reduction_ratio": 4}
    if not log_path.exists():
        print(f"  WARNING: {log_path.name} not found — using default config")
        return cfg
    for line in log_path.read_text(encoding="utf-8").splitlines():
        if re.match(r"\s*BACKBONE\s*:", line):
            cfg["backbone"] = line.split(":", 1)[1].strip().split()[0]
        elif re.match(r"\s*USE_CHANNEL_ATTENTION\s*:", line):
            cfg["use_channel_attention"] = line.split(":", 1)[1].strip().split()[0].lower() == "true"
    return cfg


def build_subset_map(datasets_dir: Path) -> dict:
    """Return Identifier -> subset dict from train/val/test CSVs."""
    mapping = {}
    for split in ("train", "val", "test"):
        csv_path = datasets_dir / f"{split}.csv"
        if not csv_path.exists():
            print(f"  WARNING: {csv_path.name} not found — '{split}' rows will be None")
            continue
        df = pd.read_csv(csv_path, usecols=["filepath"])
        for fp in df["filepath"]:
            mapping[Path(str(fp)).stem] = split
    return mapping


def reorder_columns(headers: list) -> tuple:
    """
    Reorder FINAL headers:
      - Move Estrategia right after FJAT (before Etiqueta)
      - Remove Link_img, Img, and None columns (handled as ocumap / HD15)
    Returns (new_headers, index_map) where index_map[i] gives the original
    column index for new_headers[i].
    """
    skip = {None, "Estrategia", "Link_img", "Img"}
    new_headers = []
    index_map = []
    for i, h in enumerate(headers):
        if h in skip:
            continue
        new_headers.append(h)
        index_map.append(i)
        if h == "FJAT" and "Estrategia" in headers:
            new_headers.append("Estrategia")
            index_map.append(headers.index("Estrategia"))
    return new_headers, index_map


def main():
    # --- Locate experiment and model ---
    exp_folder = Path(__file__).parent / "experiments" / EXPERIMENT_NAME
    if not exp_folder.is_dir():
        print(f"ERROR: experiment folder not found: {exp_folder}")
        sys.exit(1)

    model_path = exp_folder / "best_model.pth"
    if not model_path.exists():
        print(f"ERROR: best_model.pth not found in {exp_folder}")
        sys.exit(1)

    # --- Parse config & load model ---
    cfg             = parse_log_config(exp_folder / "training_log.txt")
    backbone        = cfg["backbone"]
    use_ca          = cfg["use_channel_attention"]
    reduction_ratio = cfg["attention_reduction_ratio"]
    print(f"Config  : backbone={backbone}  channel_attention={use_ca}")

    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    print(f"Device  : {device}")
    print(f"Loading : {model_path}")

    ckpt  = torch.load(model_path, map_location=device)
    model = build_inference_model(backbone, use_ca, reduction_ratio)
    model.load_state_dict(ckpt["model_state_dict"])
    model.to(device).eval()
    print(f"  Loaded checkpoint from epoch {ckpt.get('epoch', '?')} "
          f"(val_loss={ckpt.get('val_loss', float('nan')):.4f})")

    transform = rt.build_transforms(IMG_HEIGHT, IMG_WIDTH)

    # --- Build subset map ---
    print(f"\nBuilding subset map from: {DATASETS_DIR}")
    subset_map = build_subset_map(DATASETS_DIR)
    print(f"  {len(subset_map)} filepaths indexed")

    # --- Read FINAL sheet ---
    print(f"\nReading : {LABELS_XLSX}")
    wb = load_workbook(LABELS_XLSX, data_only=False)

    if FINAL_SHEET not in wb.sheetnames:
        print(f"ERROR: sheet '{FINAL_SHEET}' not found.")
        sys.exit(1)

    ws_final = wb[FINAL_SHEET]
    final_headers = [c.value for c in next(ws_final.iter_rows(min_row=1, max_row=1))]

    for col in ("Identifier", "Etiqueta"):
        if col not in final_headers:
            print(f"ERROR: column '{col}' not found in {FINAL_SHEET}.")
            sys.exit(1)

    id_idx  = final_headers.index("Identifier")
    et_idx  = final_headers.index("Etiqueta")
    img_idx = final_headers.index("Img") if "Img" in final_headers else None
    li_idx  = final_headers.index("Link_img") if "Link_img" in final_headers else None

    # --- Reorder columns ---
    base_headers, col_map = reorder_columns(final_headers)
    et_new_pos  = base_headers.index("Etiqueta")
    error_pos   = et_new_pos + 2  # index of "Error" in each output row

    # Output headers: base[..Etiqueta] + Pred + Error + base[after Etiqueta] + subset + ocumap + HD15
    out_headers = (
        base_headers[:et_new_pos + 1]
        + ["Pred", "Error"]
        + base_headers[et_new_pos + 1:]
        + ["subset", "ocumap", "HD15"]
    )

    # --- Run inference ---
    print("Inference ...")
    all_rows    = list(ws_final.iter_rows(min_row=2, values_only=True))
    n_total     = len(all_rows)
    result_rows = []
    n_skipped   = 0

    with torch.no_grad():
        for i, row in enumerate(all_rows, 1):
            if i % 50 == 0 or i == n_total:
                print(f"  {i}/{n_total}", end="\r", flush=True)

            identifier = row[id_idx]
            if identifier is None:
                n_skipped += 1
                continue

            # Find matching images on disk
            matches = sorted(DATA_ROOT.rglob(f"{identifier}.png"))
            if not matches:
                print(f"\n  WARNING: no image found for '{identifier}'")
                n_skipped += 1
                continue

            etiqueta_val = row[et_idx]
            true_label   = int(round(float(etiqueta_val))) if etiqueta_val is not None else None
            img_name     = row[img_idx] if img_idx is not None else None
            subset       = subset_map.get(str(identifier), None)
            link_img_val = row[li_idx] if li_idx is not None else None
            m            = re.search(r'"([^"]+)"', str(link_img_val)) if link_img_val else None
            img_prefix   = m.group(1) if m else ""

            for img_path in matches:
                rel_path = str(img_path.relative_to(DATA_ROOT))

                with Image.open(img_path) as img:
                    tensor = transform(img.convert("RGB")).unsqueeze(0).to(device)

                pred_raw   = round(float(torch.clamp(model(tensor).squeeze(), 1.0, 5.0).item()), 3)
                pred_label = int(round(pred_raw))
                error      = abs(pred_label - true_label) if true_label is not None else None

                # Build output row
                base_vals = [row[j] for j in col_map]
                out_row = (
                    base_vals[:et_new_pos + 1]
                    + [pred_raw, error]
                    + base_vals[et_new_pos + 1:]
                    + [subset, img_name, rel_path, img_prefix]  # img_prefix: extra, not a sheet column
                )
                result_rows.append(out_row)

    print(f"\n  Done: {len(result_rows)} processed, {n_skipped} skipped.")

    # Sort by error descending (worst mismatches first)
    result_rows.sort(key=lambda r: (r[error_pos] is None, -(r[error_pos] or 0)))

    # --- Build output sheet ---
    ts_match   = re.search(r"(\d{8}_\d{6})", exp_folder.name)
    sheet_name = f"PRED_{ts_match.group(1)}" if ts_match else f"PRED_{exp_folder.name[:20]}"

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws_pred = wb.create_sheet(sheet_name)
    ws_pred.append(out_headers)

    ocumap_col = out_headers.index("ocumap") + 1  # 1-based
    hd15_col   = out_headers.index("HD15") + 1     # 1-based

    for row_num, row_data in enumerate(result_rows, start=2):
        img_prefix_val = row_data[-1]   # extra element — not a sheet column
        img_name_val   = row_data[-3]
        filepath_val   = row_data[-2]

        write_row = list(row_data[:-1])  # drop img_prefix — not written to sheet
        write_row[-2] = None  # ocumap — replaced by hyperlink
        write_row[-1] = None  # HD15   — replaced by hyperlink
        ws_pred.append(write_row)

        if img_name_val:
            ws_pred.cell(row=row_num, column=ocumap_col).value = (
                f'=HYPERLINK("{img_prefix_val}{img_name_val}", "Ver Imagen")'
            )
        if filepath_val:
            ws_pred.cell(row=row_num, column=hd15_col).value = (
                f'=HYPERLINK("{filepath_val}", "Ver Imagen")'
            )

    wb.save(LABELS_XLSX)

    # --- Summary ---
    n_correct = sum(1 for r in result_rows if r[error_pos] == 0)
    n_wrong   = sum(1 for r in result_rows if r[error_pos] is not None and r[error_pos] > 0)
    by_err    = {}
    for r in result_rows:
        e = r[error_pos]
        if e is not None:
            by_err[e] = by_err.get(e, 0) + 1

    print(f"\nSheet '{sheet_name}' saved to {LABELS_XLSX}")
    print(f"  Total  : {len(result_rows)}")
    print(f"  Correct: {n_correct}  ({100 * n_correct / max(len(result_rows), 1):.1f}%)")
    print(f"  Wrong  : {n_wrong}")
    print("  Error distribution:")
    for err in sorted(by_err):
        print(f"    |error|={err} : {by_err[err]} samples")


if __name__ == "__main__":
    main()
