"""
Shared helpers for reading SOTER labels from ``labels.xlsx`` and joining
them with INSPECTROAD per-segment metrics from ``roads_qualified/all_roads.csv``.

Used by the visualisation scripts (heatmap, prediction shapefiles) and by
any analysis that needs SOTER labels + INSPECTROAD defect counts aligned
on segment identifier.
"""

from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook


def get_last_pred_sheet(xlsx_path: Path) -> str:
    """Return the name of the last sheet whose name starts with 'PRED_'."""
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    pred_sheets = [s for s in wb.sheetnames if s.startswith("PRED_")]
    wb.close()
    if not pred_sheets:
        raise ValueError(f"No PRED_* sheet found in {xlsx_path}")
    return pred_sheets[-1]


def build_attribute_table(xlsx_path: Path, sheet_name: str | None = None) -> pd.DataFrame:
    """Build the SOTER attribute table from a PRED sheet.

    Columns: identifier, label, raw_est, estimate, subset.
    If *sheet_name* is None, the latest PRED_* sheet is used.
    """
    if sheet_name is None:
        sheet_name = get_last_pred_sheet(xlsx_path)
    df = pd.read_excel(xlsx_path, sheet_name=sheet_name)
    attrs = pd.DataFrame({
        "identifier": df["Identifier"],
        "label":      df["Etiqueta"],
        "raw_est":    df["Pred"],
        "estimate":   np.round(df["Pred"]).astype(int),
        "subset":     df["subset"],
    })
    return attrs.drop_duplicates(subset="identifier", keep="first")


def load_calibration_probs(exp_dir: Path) -> pd.DataFrame | None:
    """Load Gaussian and Ordinal calibration CSVs, return merged DataFrame.

    Columns: identifier, gau_p1..5, gau_conf, gau_pred, ord_p1..5, ord_conf, ord_pred.
    Returns None if no calibration CSVs with identifier column are found.
    """
    frames = []
    for cal_type, prefix in [("gaussian", "gau"), ("ordinal", "ord")]:
        parts = []
        for split in ("train", "val", "test"):
            csv_path = exp_dir / f"calibration_{cal_type}_{split}.csv"
            if csv_path.exists():
                df = pd.read_csv(csv_path)
                if "identifier" not in df.columns:
                    continue
                parts.append(df)
        if not parts:
            continue
        df_all = pd.concat(parts, ignore_index=True)
        df_all = df_all.drop_duplicates(subset="identifier", keep="first")
        rename = {f"prob_{k}": f"{prefix}_p{k}" for k in range(1, 6)}
        df_all = df_all.rename(columns=rename)
        prob_cols = [f"{prefix}_p{k}" for k in range(1, 6)]
        df_all[f"{prefix}_conf"] = df_all[prob_cols].max(axis=1)
        df_all[f"{prefix}_pred"] = df_all["pred_class"]
        keep_cols = ["identifier"] + prob_cols + [f"{prefix}_conf", f"{prefix}_pred"]
        frames.append(df_all[keep_cols])

    if not frames:
        return None
    result = frames[0]
    for f in frames[1:]:
        result = result.merge(f, on="identifier", how="outer")
    return result


def load_soter_inspectroad(
    labels_xlsx: Path,
    all_roads_csv: Path,
    how: str = "left",
    sheet_name: str | None = None,
) -> pd.DataFrame:
    """Return SOTER attribute table merged with INSPECTROAD metrics on ``identifier``."""
    attrs = build_attribute_table(labels_xlsx, sheet_name)
    roads = pd.read_csv(all_roads_csv)
    return attrs.merge(roads, on="identifier", how=how)
